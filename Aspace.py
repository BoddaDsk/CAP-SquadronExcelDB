import openpyxl
import main
wb = openpyxl.load_workbook(main.filename)
ws = wb.active
def aspaceTest_check(num_cadets):
    AspaceTestIndex = []
    row_num = 1
    for row in ws.iter_rows(min_row=2, min_col=15, max_col=15, max_row=num_cadets, values_only=True):
        row_num +=1
        if row[0] > 79:
            AspaceTestIndex.append(row_num)
    return AspaceTestIndex

def aspaceInt_check(num_cadets):
    AspaceIntIndex = []
    row_num = 1
    for row in ws.iter_rows(min_row=2, min_col=35, max_col=35, max_row=num_cadets, values_only=True):
        row_num +=1
        if row[0] != "N/A" and row[0] != "None":
            AspaceIntIndex.append(row_num)
    return AspaceIntIndex

def aspacePass(num_cadets):
    aspacePass = set(aspaceInt_check(num_cadets) + aspaceTest_check(num_cadets))
    return aspacePass