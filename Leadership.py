import openpyxl
import main
wb = openpyxl.load_workbook(main.filename)
ws = wb.active
def leadTest_check(num_cadets):
    LeadTestIndex = []
    row_num = 1
    for row in ws.iter_rows(min_row=2, min_col=13, max_col=13, max_row=num_cadets, values_only=True):
        row_num +=1
        if row[0] > 79:
            LeadTestIndex.append(row_num)
    return LeadTestIndex

def leadInt_check(num_cadets):
    LeadIntIndex = []
    row_num = 1
    for row in ws.iter_rows(min_row=2, min_col=37, max_col=37, max_row=num_cadets, values_only=True):
        row_num +=1
        if row[0] != "None":
            LeadIntIndex.append(row_num)
    return LeadIntIndex

def leadPass(num_cadets):
    leadPass = set(leadTest_check(num_cadets) + leadInt_check(num_cadets))
    return leadPass








