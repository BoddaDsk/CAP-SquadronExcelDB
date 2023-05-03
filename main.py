import openpyxl

day = input("\nPlease input the day the Excel file was generated [No leading 0s]: ")
month = input('Please input the number of month this was generated [No leading 0s]: ')
year = input("Input the last 2 digits of the year the file was generated: ")
filename = f'Cadet_Full_Track_Report-{month}_{day}_20{year}.xlsx'
wb = openpyxl.load_workbook(filename)
ws = wb.active
start = 0
print('\n<------------------------------------------------------------------------>')
print('Welcome to the Jimmy Stewart Composite Squadron Cadet Programs Data Base')

#-----------------LEADERSHIP---------------------------------------------------------------------------------------------
def leadTest_check(num_cadets):
    LeadTestIndex = []
    row_num = 1
    for row in ws.iter_rows(min_row=2, min_col=13, max_col=13, max_row=num_cadets, values_only=True):
        row_num +=1
        if row[0] > 79:
            LeadTestIndex.append(row_num)
    return LeadTestIndex

def leadInt_check(num_cadets):
    LeadIntIndex = []
    row_num = 1
    for row in ws.iter_rows(min_row=2, min_col=37, max_col=37, max_row=num_cadets, values_only=True):
        row_num +=1
        if row[0] != "None":
            LeadIntIndex.append(row_num)
    return LeadIntIndex

def leadPass(num_cadets):
    LeadPass = set(leadTest_check(num_cadets) + leadInt_check(num_cadets))
    return LeadPass
#-----------------AEROSPACE---------------------------------------------------------------------------------------------
def aspaceTest_check(num_cadets):
    AspaceTestIndex = []
    row_num = 1
    for row in ws.iter_rows(min_row=2, min_col=15, max_col=15, max_row=num_cadets, values_only=True):
        row_num +=1
        if row[0] > 79:
            AspaceTestIndex.append(row_num)
    return AspaceTestIndex

def aspaceInt_check(num_cadets):
    AspaceIntIndex = []
    row_num = 1
    for row in ws.iter_rows(min_row=2, min_col=35, max_col=35, max_row=num_cadets, values_only=True):
        row_num +=1
        if row[0] != "N/A" and row[0] != "None":
            AspaceIntIndex.append(row_num)
    return AspaceIntIndex

def aspacePass(num_cadets):
    aspacePass = set(aspaceInt_check(num_cadets) + aspaceTest_check(num_cadets))
    return aspacePass
#--------------------------------------------------------------------------------------------------------------------------

while start < 1:
    print('\nSelect which of these you would like to run: ')
    print('1. List cadets with completed leadership')
    print('2. List of cadets with completed aerospace')
    print('3. List of cadets needing a drill test')
    print('4. Exit the program')
    menu_start = int(input("Your Selection: "))

    if menu_start == 4:
        print("The program will now end.")
        break

    num_cadets = int(input('How many cadets are there? \n'))
   

    if menu_start == 1:
        for x in leadPass(num_cadets):
            print(ws[f'C{x}'].value + " " + ws[f'B{x}'].value)
        print("\nWould you like to continue?")
        cont = str(input("(Y/N): "))

    if menu_start == 2:
        for x in aspacePass(num_cadets):
            print(ws[f'C{x}'].value + " " + ws[f'B{x}'].value)
        print("\nWould you like to continue?")
        cont = str(input("(Y/N): "))

    if menu_start == 3:
        drTest = []
        for x in leadPass(num_cadets):
            for y in aspacePass(num_cadets):
                if x==y:
                    drTest.append(x)
        for z in drTest:
            print(ws[f'C{z}'].value + " " + ws[f'B{z}'].value)
        print("\nWould you like to continue?")
        cont = str(input("(Y/N): "))
        print(cont)

    if cont == 'N' or cont == 'n':
        print("The program will now end.")
        start = 1

print('\n<------------------------------------------------------------------------>')
